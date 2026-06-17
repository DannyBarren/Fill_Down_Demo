"""Export the filled-down results back to Excel or CSV.

Strategies:
    * ``export_preserving_original`` – re-open the *original* uploaded workbook
      with openpyxl and write only the ``New Account`` column. This keeps all
      original formatting, formulas, column widths, column order and other
      sheets intact, and (optionally) adds a clean "Fill Down Summary" sheet.
    * ``export_dataframe`` – write a clean workbook from scratch (used for CSV
      uploads or when the original bytes are unavailable), preserving column
      order and optionally adding the summary sheet.
    * ``export_csv`` – a tidy CSV with internal columns stripped and
      a derived base-account column, ready to import into accounting/ERP tools.
"""

from __future__ import annotations

import io
from typing import List, Optional

import pandas as pd

from src.config import Config
from src.data_loader import SIM_TEXT_COL
from utils.account_codes import base_account, normalize_code
from utils.logging_setup import get_logger

logger = get_logger(__name__)

# Internal helper columns that must never leak into an export. In addition to
# this explicit set, every column whose name starts with "_" is treated as
# internal (e.g. ``_sim_text``, ``_base_sig`` and the spreadsheet meta columns
# ``_confidence`` / ``_engine`` / ``_action`` / ``_why`` / ``_select``). Real
# client columns never start with an underscore.
_INTERNAL_COLS = {SIM_TEXT_COL}

SUMMARY_SHEET = "Fill Down Summary"


def _clean_df(df: pd.DataFrame) -> pd.DataFrame:
    """Drop internal helper columns while preserving the original column order."""
    drop = [c for c in df.columns
            if c in _INTERNAL_COLS or str(c).startswith("_")]
    return df.drop(columns=drop, errors="ignore")


def export_dataframe(
    df: pd.DataFrame,
    fmt: str = "xlsx",
    sheet_name: str = "Transactions",
    summary: Optional[pd.DataFrame] = None,
) -> bytes:
    """Serialise the dataframe to xlsx or csv bytes (internal cols stripped)."""
    clean = _clean_df(df)
    buffer = io.BytesIO()
    if fmt == "csv":
        clean.to_csv(buffer, index=False)
    else:
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            clean.to_excel(writer, index=False, sheet_name=sheet_name)
            _autosize_columns(writer.sheets[sheet_name], clean)
            if summary is not None and not summary.empty:
                summary.to_excel(writer, index=False, sheet_name=SUMMARY_SHEET)
                _autosize_columns(writer.sheets[SUMMARY_SHEET], summary)
    buffer.seek(0)
    logger.info("export_dataframe", fmt=fmt, rows=len(clean),
                summary=summary is not None)
    return buffer.getvalue()


def export_csv(
    df: pd.DataFrame,
    na_col: str,
    add_base_column: bool = True,
) -> bytes:
    """Return a clean CSV with normalised codes, ready for accounting import."""
    clean = _clean_df(df).copy()
    if na_col in clean.columns:
        clean[na_col] = clean[na_col].fillna("").map(
            lambda v: normalize_code(v) if str(v).strip() else "")
        if add_base_column:
            base_col = f"{na_col} (Base)"
            insert_at = clean.columns.get_loc(na_col) + 1
            clean.insert(insert_at, base_col,
                         clean[na_col].map(lambda v: base_account(v)))
    buffer = io.BytesIO()
    clean.to_csv(buffer, index=False)
    buffer.seek(0)
    logger.info("export_csv", rows=len(clean))
    return buffer.getvalue()


def export_preserving_original(
    original_bytes: bytes,
    df: pd.DataFrame,
    na_col: str,
    config: Config,
    sheet_name=0,
    header_row: int = 1,
    extra_header_candidates: Optional[List[str]] = None,
    summary: Optional[pd.DataFrame] = None,
) -> bytes:
    """Write ``New Account`` values into a copy of the original workbook.

    Preserves the original column order, formatting and other sheets. Optionally
    appends a "Fill Down Summary" sheet. Falls back to :func:`export_dataframe`
    if the original cannot be opened.
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(original_bytes))
        ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else wb[sheet_name]

        candidates = {c.strip().lower() for c in config.columns.new_account}
        candidates.add(na_col.strip().lower())
        for extra in extra_header_candidates or []:
            candidates.add(str(extra).strip().lower())

        target_col_idx: Optional[int] = None
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            value = ws.cell(row=header_row, column=col).value
            if value is not None and str(value).strip().lower() in candidates:
                target_col_idx = col
                break

        if target_col_idx is None:
            target_col_idx = max_col + 1
            ws.cell(row=header_row, column=target_col_idx, value=na_col)
            logger.info("export_added_new_account_column", col=target_col_idx)

        values: List[str] = df[na_col].fillna("").astype(str).tolist()
        for offset, val in enumerate(values):
            ws.cell(row=header_row + 1 + offset, column=target_col_idx,
                    value=(val if val.strip() else None))

        if summary is not None and not summary.empty:
            _write_summary_sheet(wb, summary)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info("export_preserving_original", rows=len(values),
                    col=target_col_idx, summary=summary is not None)
        return buffer.getvalue()
    except Exception as exc:  # noqa: BLE001
        logger.warning("preserve_export_failed_fallback", error=str(exc))
        return export_dataframe(df, fmt="xlsx", summary=summary)


def _write_summary_sheet(wb, summary: pd.DataFrame) -> None:
    """Add (or replace) the summary sheet in an openpyxl workbook."""
    from openpyxl.styles import Font

    if SUMMARY_SHEET in wb.sheetnames:
        del wb[SUMMARY_SHEET]
    ws = wb.create_sheet(SUMMARY_SHEET)
    ws.append(list(summary.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for _, row in summary.iterrows():
        ws.append([row[c] for c in summary.columns])
    # Reasonable widths.
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28


def _autosize_columns(worksheet, df: pd.DataFrame, max_width: int = 60) -> None:
    """Best-effort column auto-sizing for fresh exports."""
    from openpyxl.utils import get_column_letter

    for idx, col in enumerate(df.columns, start=1):
        series = df[col].astype(str)
        width = max([len(str(col))] + [len(v) for v in series.head(200)] + [4])
        worksheet.column_dimensions[get_column_letter(idx)].width = min(
            width + 2, max_width
        )
